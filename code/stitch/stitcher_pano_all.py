import os.path as osp
import os
import cv2
import numpy as np
import openpyxl

import torch
from sg.matching import Matching
from sg.utils import (AverageTimer, VideoStreamer,
                          make_matching_plot_fast, frame2tensor)
from loftr.loftr import LoFTR, default_cfg

import argparse

torch.set_grad_enabled(False)
cv2.ocl.setUseOpenCL(False)
W = 640
H = 480



def cv_imread(file_path):
    cv_img = cv2.imdecode(np.fromfile(file_path, dtype=np.uint8), -1)
    return cv_img

device = 'cuda' if torch.cuda.is_available() else 'cpu'
print('Running inference on device \"{}\"'.format(device))
config = {
    'superpoint': {
        'nms_radius': 8,
        'keypoint_threshold': 0.001,
        'max_keypoints': -1
    },
    'superglue': {
        'weights': 'outdoor',
        'sinkhorn_iterations': 100,
        'match_threshold': 0.1,
    }
}
matching = Matching(config).eval().to(device)
matcher = LoFTR(config=default_cfg)
image_type = 'outdoor'
if image_type == 'indoor':
    matcher.load_state_dict(torch.load("loftr/weights/indoor_ds_new.ckpt")['state_dict'])
elif image_type == 'outdoor':
    matcher.load_state_dict(torch.load("loftr/weights/outdoor_ds.ckpt")['state_dict'])
else:
    raise ValueError("Wrong image_type is given.")
matcher = matcher.eval().to(device)


def nn_match_two_way(desc1, desc2, nn_thresh=0.7):
    if desc1.shape[1] == 0 or desc2.shape[1] == 0:
        return np.zeros((3, 0))
    if nn_thresh < 0.0:
        raise ValueError('\'nn_thresh\' should be non-negative')
    # Compute L2 distance. Easy since vectors are unit normalized.
    dmat = np.dot(desc1, desc2.T)
    dmat = np.sqrt(2 - 2 * np.clip(dmat, -1, 1))
    # Get NN indices and scores.
    idx = np.argmin(dmat, axis=1)
    scores = dmat[np.arange(dmat.shape[0]), idx]
    # Threshold the NN matches.
    keep = scores < nn_thresh
    # Check if nearest neighbor goes both directions and keep those.
    idx2 = np.argmin(dmat, axis=0)
    keep_bi = np.arange(len(idx)) == idx2[idx]
    keep = np.logical_and(keep, keep_bi)
    idx = idx[keep]
    scores = scores[keep]
    # Get the surviving point indices.
    m_idx1 = np.arange(desc1.shape[0])[keep]
    m_idx2 = idx
    # Populate the final 3xN match data structure.
    matches = np.zeros((3, int(keep.sum())))
    matches[0, :] = m_idx1
    matches[1, :] = m_idx2
    matches[2, :] = scores
    return matches


def sp_stitch_two(im11, im22):
    # print(im11.shape, im22.shape)
    im11 = cv2.cvtColor(im11, cv2.COLOR_BGR2GRAY)
    im22 = cv2.cvtColor(im22, cv2.COLOR_BGR2GRAY)
    im11 = cv2.resize(im11.astype('float32'), (W, H))
    im22 = cv2.resize(im22.astype('float32'), (W, H))

    inp0 = frame2tensor(im11, device)
    inp1 = frame2tensor(im22, device)
    pred, pred0, pred1 = matching({'image0': inp0, 'image1': inp1})
    # pred, pred0, pred1 = matching({'image0': inp0, 'image1': inp1, 'keypoints0': x1, 'keypoints1': x2, 'descriptors0':des1, 'descriptors1':des2})

    pred = {k: v[0].cpu().numpy() for k, v in pred.items()}
    kpts0, kpts1 = pred['keypoints0'], pred['keypoints1']
    # mdesc0, mdesc1 = pred['descriptors0'], pred['descriptors1']
    des_sp0, des_sp1 = (pred0['descriptors'][0].t()).cpu().numpy(), (pred1['descriptors'][0].t()).cpu().numpy()
    sp_matches = nn_match_two_way(des_sp0, des_sp1, nn_thresh=0.8)
    better = sp_matches[0:2, :]
    # for m, n, scores in sp_matches:
    #     better.append((m, n))
    # matches = pred['matches0']
    # confidence = pred['matching_scores0']
    _, num_match = better.shape
    mkpts0 = np.float32([kpts0[int(i)] for i in better[0, :]])
    mkpts1 = np.float32([kpts1[int(i)] for i in better[1, :]])
    # valid = matches > -1
    # mkpts0 = kpts0[valid]
    # mkpts1 = kpts1[matches[valid]]

    print(len(mkpts0), len(mkpts1))
    if num_match > 10:
        (HH, status) = cv2.findHomography(mkpts1, mkpts0, cv2.RANSAC, ransacReprojThreshold=5.0)
        print(HH)
        if HH[0][0] <= 3 and HH[0][0] >= 0.4 and HH[1][1] <= 3 and HH[1][1] >= 0.4 and abs(HH[0][-1]) < W and abs(
                HH[1][-1]) < H:
            return True, HH, num_match
        return False, None, num_match
    else: return False, None, num_match


def sg_stitch_two(im11, im22):
    # print(im11.shape, im22.shape)
    im11 = cv2.cvtColor(im11, cv2.COLOR_BGR2GRAY)
    im22 = cv2.cvtColor(im22, cv2.COLOR_BGR2GRAY)
    im11 = cv2.resize(im11.astype('float32'), (W, H))
    im22 = cv2.resize(im22.astype('float32'), (W, H))

    inp0 = frame2tensor(im11, device)
    inp1 = frame2tensor(im22, device)
    # image0, inp0, scales0 = read_image(
    #     im11, device, [W*i, H], 0, None)
    # image1, inp1, scales1 = read_image(
    #     im22, device, [W, H], 0, None)
    pred, pred0, pred1 = matching({'image0': inp0, 'image1': inp1})
    # pred, pred0, pred1 = matching({'image0': inp0, 'image1': inp1, 'keypoints0': x1, 'keypoints1': x2, 'descriptors0':des1, 'descriptors1':des2})

    pred = {k: v[0].cpu().numpy() for k, v in pred.items()}
    kpts0, kpts1 = pred['keypoints0'], pred['keypoints1']
    mdesc0, mdesc1 = pred['descriptors0'], pred['descriptors1']
    matches = pred['matches0']
    confidence = pred['matching_scores0']
    valid = matches > -1
    mkpts0 = kpts0[valid]
    mkpts1 = kpts1[matches[valid]]

    print(len(mkpts0), len(mkpts1))
    if len(mkpts0) > 10:
        (HH, status) = cv2.findHomography(mkpts1, mkpts0, cv2.RANSAC, ransacReprojThreshold=5.0)
        # result = cv2.warpPerspective(im22, HH, (int(im11.shape[1] * 1.1), im11.shape[0]))
        # result[0:im11.shape[0], 0:im11.shape[1]] = im11
        print(HH)
        return True, HH, len(mkpts0)
    else: return False, None, 0
    # cv2.imshow('result1', result)
    # 将图片B传入result图片最左端

def loftr_stitch_two(im11, im22):
    # print(im11.shape, im22.shape)
    im11 = cv2.cvtColor(im11, cv2.COLOR_BGR2GRAY)
    im22 = cv2.cvtColor(im22, cv2.COLOR_BGR2GRAY)
    im11 = cv2.resize(im11.astype('float32'), (W, H))
    im22 = cv2.resize(im22.astype('float32'), (W, H))

    inp0 = frame2tensor(im11, device)
    inp1 = frame2tensor(im22, device)

    batch = {'image0': inp0, 'image1': inp1}
    with torch.no_grad():
        matcher(batch)
        mkpts0 = batch['mkpts0_f'].cpu().numpy()
        mkpts1 = batch['mkpts1_f'].cpu().numpy()
        mconf = batch['mconf'].cpu().numpy()
    print(len(mkpts0), len(mkpts1))
    if len(mkpts0) > 10:
        (HH, status) = cv2.findHomography(mkpts1, mkpts0, cv2.RANSAC, ransacReprojThreshold=5.0)
        print(HH)
        if HH[0][0] <= 3 and HH[0][0] >= 0.35 and HH[1][1] <= 3 and HH[1][1] >= 0.35 and abs(HH[0][-1]) < W and abs(HH[1][-1]) < H:
            return True, HH, len(mkpts0)
        else:
            return False, None, len(mkpts0)
    else: return False, None, 0


def sift_stitch_two(im11, im22):
    im11 = cv2.cvtColor(im11, cv2.COLOR_BGR2GRAY)
    im22 = cv2.cvtColor(im22, cv2.COLOR_BGR2GRAY)
    im11 = cv2.resize(im11.astype('uint8'), (W, H))
    im22 = cv2.resize(im22.astype('uint8'), (W, H))

    sift = cv2.SIFT_create()
    kp1, des1 = sift.detectAndCompute(im11, None)  # des是描述子
    kp2, des2 = sift.detectAndCompute(im22, None)  # des是描述子
    kp1 = np.float32([kp.pt for kp in kp1])
    kp2 = np.float32([kp.pt for kp in kp2])
    # print(des1, des2)
    if des1 is None or des2 is None:
        return False, None, 0
    matches = cv2.BFMatcher().knnMatch(des1, des2, k=2)
    good = []

    for m in matches:
        # 当最近距离跟次近距离的比值小于ratio值时，保留此匹配对
        if len(m) == 2 and m[0].distance < m[1].distance * 0.7:
            # 存储两个点在featuresA, featuresB中的索引值
            good.append((m[0].trainIdx, m[0].queryIdx))

    num_match = len(good)
    ptsA = np.float32([kp1[i] for (_, i) in good])
    ptsB = np.float32([kp2[i] for (i, _) in good])
    if num_match > 10:
        (HH, status) = cv2.findHomography(ptsB, ptsA, cv2.RANSAC, ransacReprojThreshold=5.0)

        print(HH)
        if HH[0][0] <= 2 and HH[0][0] >= 0.5 and HH[1][1] <= 2 and HH[1][1] >= 0.5 and abs(HH[0][-1]) < W and abs(HH[1][-1]) < H:
            return True, HH, num_match
        else:
            return False, None, num_match
    else:
        return False, None, num_match



def match_ratio(name1, name2):
    cnt = 0
    for i in range(len(name1)):
        if name1[i] != name2[i]:
            cnt += 1
    return cnt


def hh_stitch(image_list, HH_list, match_method):
    # print(len(image_list), len(HH_list))
    cur_index = len(HH_list)
    for i in range(cur_index + 1, len(image_list)):

        if match_method == 'sift':
            status, HH, num_match = sift_stitch_two(image_list[i - 1], image_list[i])
        if match_method == 'sp':
            status, HH, num_match = sp_stitch_two(image_list[i - 1], image_list[i])
        if match_method == 'sg':
            status, HH, num_match = sg_stitch_two(image_list[i - 1], image_list[i])
        if match_method == 'loftr':
            status, HH, num_match = loftr_stitch_two(image_list[i - 1], image_list[i])

        if not status:
            return 1, [], []
        HH_list.append(HH)

    return 0, image_list, HH_list


def treat_list(image_list, homo_list):
    center = len(image_list) // 2
    HH_bias = np.array([[1, 0, 0], [0, 1, 0], [0, 0, 1]])
    for i in range(0, center):
        HH_bias = np.dot(HH_bias, homo_list[i])
    HH_bias = np.linalg.inv(HH_bias)
    HH_cur_to_first = np.array([[1, 0, 0], [0, 1, 0], [0, 0, 1]])

    HH_use_list = []
    min_w = 1000
    min_h = 1000
    max_w = -1000
    max_h = -1000
    for i in range(len(image_list)):
        HH_cur = np.dot(HH_bias, HH_cur_to_first)
        f1 = np.dot(HH_cur, np.array([0, 0, 1]))
        f1 = f1 / f1[-1]
        f2 = np.dot(HH_cur, np.array([W, 0, 1]))
        f2 = f2 / f2[-1]
        f3 = np.dot(HH_cur, np.array([0, H, 1]))
        f3 = f3 / f3[-1]
        f4 = np.dot(HH_cur, np.array([W, H, 1]))
        f4 = f4 / f4[-1]
        print(i, HH_cur)
        HH_use = HH_cur
        HH_use_list.append(HH_use)
        min_w = min(min_w, min([f1[0], f2[0], f3[0], f4[0]]))
        min_h = min(min_h, min([f1[1], f2[1], f3[1], f4[1]]))
        max_w = max(max_w, max([f1[0], f2[0], f3[0], f4[0]]))
        max_h = max(max_h, max([f1[1], f2[1], f3[1], f4[1]]))
        if i < len(image_list) - 1:
            HH_cur_to_first = np.dot(HH_cur_to_first, homo_list[i])
    max_w_int = int(max_w + 1)
    max_h_int = int(max_h + 1)
    min_w_int = int(min_w - 1)
    min_h_int = int(min_h - 1)
    # print(max_w_int, max_h_int, min_w_int, min_h_int)
    newW = - min_w_int + max_w_int
    newH = - min_h_int + max_h_int
    # print(newW, newH)
    pano = np.zeros([newH, newW, 3], np.uint8)
    for i in range(len(image_list)):
        HH_use = HH_use_list[i]
        HH_use[0][-1] += - min_w_int
        HH_use[1][-1] += - min_h_int
        image_cur = cv2.warpPerspective(image_list[i], HH_use, (newW, newH))
        mask = image_cur != 0
        pano[mask] = image_cur[mask]

        if i < len(image_list) - 1:
            HH_cur_to_first = np.dot(HH_cur_to_first, homo_list[i])

    return pano


if __name__ == "__main__":

    parser = argparse.ArgumentParser()
    parser.add_argument('--dataset', default='kjg1', help='kjg1, kjg2, kjg3, kjg4, hw')
    parser.add_argument('--method', default='loftr', help='sift, sp, sg, loftr')
    parser.add_argument('--root', default='/home/fsh/stitching_and_registration/input_data/stitch_data/stitch')
    parser.add_argument('--outfile', default='/home/fsh/stitching_and_registration/output_data/stitch_out')
    args = parser.parse_args()

    dataset = args.dataset
    match_method = args.method
    root = args.root
    out_root = args.outfile

    out_path = osp.join(out_root, dataset)

    if not osp.exists(out_path):
        os.makedirs(out_path)
    img_path = []

    if dataset == 'kjg1':
        base_path = osp.join(root, '(例)泉城变电站白天例行任务20220306070002')
        excel_name = '(例)泉城变电站白天例行任务20220306070002.xlsx'
        multi_list = [[129, 131, 133], [141, 142, 144, 146], [154, 156, 158], [172, 174, 175], [178, 179, 181],
                      [200, 201, 202, 203, 205, 207], [210, 212, 214, 217, 218], [229, 231, 233], [242, 243, 245, 246],
                      [308, 309, 311, 313], [321, 322, 323, 324, 326, 328]]
        start_line = 99
        image_col = 9
        name_col = 1
        part_col = 2
    if dataset == 'kjg2':
        base_path = osp.join(root, '(例)泉城变电站白天例行任务20220321070002')
        excel_name = '(例)泉城变电站白天例行任务20220321070002.xlsx'
        multi_list = [[129, 130, 131], [162, 164, 166, 168, 170], [202, 204, 205], [231, 232, 233, 234, 236, 238], [241, 243, 245, 248, 249, 251],
                      [288, 290, 292], [297, 299, 301], [302, 305]]
        start_line = 129
        image_col = 9
        name_col = 1
        part_col = 2
    if dataset == 'kjg3':
        base_path = osp.join(root, '(例)泉城变电站白天例行任务20220428070002')
        excel_name = '(例)泉城变电站白天例行任务20220428070002.xlsx'
        multi_list = [[172, 174, 176], [190, 191, 193, 195], [244, 246, 248], [290, 291, 292],
                      [371, 372, 373, 374, 376, 378],
                      [393, 394, 396, 397], [414, 415, 418]]
        start_line = 148
        image_col = 9
        name_col = 1
        part_col = 2
    if dataset == 'kjg4':
        base_path = osp.join(root, '(例)泉城变电站白天例行任务20220501070002')
        excel_name = '(例)泉城变电站白天例行任务20220501070002.xlsx'
        multi_list = [[346, 347, 348], [438, 439, 440, 441], [164, 166, 168], [202, 203, 205, 207], [204, 206, 208, 209],
                      [262, 263, 265, 267], [385, 386, 387, 388, 390, 392], [407, 408, 409, 410]]
        start_line = 160
        image_col = 9
        name_col = 1
        part_col = 2
        # print(os.path.exists('/data/fsh/pinjie_data/(例)泉城变电站白天例行任务20220501070002/image/110kV设备区1主变1121L低抗间隔112117接地刀闸A相本体刀闸外观20220501070702.jpg'))
    if dataset == 'hw':
        base_path = osp.join(root, '室外巡检机器人红外任务20220815172558')
        excel_name = '室外巡检机器人红外任务20220815172558.xlsx'
        multi_list = [[343, 344, 345], [346, 347, 348], [372, 373, 374, 375, 376, 377, 378],
                      [443, 444, 445, 446, 447, 448, 449, 450, 451],
                      [487, 488, 489], [502, 503, 504], [573, 574, 575, 578, 579], [600, 601, 602, 603],
                      [650, 651, 652], [665, 666, 667], [522, 523, 524, 525, 526, 527, 528, 529]]
        start_line = 342
        image_col = 10
        name_col = 4
        part_col = 1

    work_book = openpyxl.load_workbook(osp.join(base_path, excel_name))
    sheet_1 = work_book.active
    # print(sheet_1.dimensions)
    row_sum = sheet_1.max_row
    row_0_value = sheet_1.cell(15, 6)
    # print(row_sum, row_0_value)
    name_last = ''
    stitcher = cv2.Stitcher.create(cv2.Stitcher_PANORAMA)
    image_list = []
    notfirst = 0
    cnt = 0
    cnt_pinjie = 0
    index_list = []
    name_list = []
    # start_line = 99
    # start_line = 129
    # start_line = 148
    # start_line = 346
    # image_col = 9
    # start_line = 342
    # # image_col = 10
    # max_row_sum = 500
    # name_col = 4

    for index, demo_list in enumerate(multi_list):
        # if index > 2:
        #     break
        print(demo_list)
        image_list = []
        index_list = []
        name_list = []
        homo_list = []
        for i in demo_list:
            name = sheet_1.cell(i, name_col).value
            url_map = sheet_1.cell(i, image_col).value
            url = sheet_1.cell(i, image_col).hyperlink.target
            image_path = osp.join(base_path+'/image', url[6:])

            img1 = cv_imread(image_path)
            img1 = cv2.resize(img1.astype('float32'), (W, H))
            img1_gray = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY)
            image_list.append(img1)
            index_list.append(i)
            name_list.append(name)
            notfirst = 1

        if len(image_list)>1:
            cnt += 1
            (status, image_list, homo_list) = hh_stitch(image_list, homo_list, match_method)
            image_name = dataset + '_' + str(demo_list[0]) + '_' + match_method + '_' + str(len(image_list)) + '.jpg'
            if status == cv2.Stitcher_OK:
                pano = treat_list(image_list, homo_list)
                cnt_pinjie += 1
                print("拼接成功.")
                image_path = osp.join(out_path, image_name)
                cv2.imwrite(image_path, pano)
                ori_dir_name = dataset + '_' + str(demo_list[0]) + '_' + str(len(image_list))
                ori_dir = osp.join(out_path, ori_dir_name)
                if not osp.exists(ori_dir):
                    os.makedirs(ori_dir)
                for j in range(len(image_list)):
                    image_path = osp.join(ori_dir, str(demo_list[0] + j) + '_.jpg')
                    cv2.imwrite(image_path, image_list[j])

            else:
                print("拼接失败.")
            name_last = name

