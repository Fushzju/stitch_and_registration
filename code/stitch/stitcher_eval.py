import os.path as osp
import os
import argparse
import cv2
import numpy as np
import openpyxl

import torch
import random

from sg.matching import Matching
from sg.utils import frame2tensor
from sg.utils import make_matching_plot_color, draw_SPMatches
from loftr.loftr import LoFTR, default_cfg
from loftr.utils.plotting import make_matching_plot_color_Loftr
from sklearn.metrics.cluster import mutual_info_score
import matplotlib.cm as cm
from scipy.stats import entropy, pearsonr

torch.set_grad_enabled(False)
W = 640
H = 480
cv2.ocl.setUseOpenCL(False)


# 打开文件方式1：

# 方式2：

def cv_imread(file_path):
    """可读取图片（路径为中文）

    :param file_path: 图片路径
    :return:
    """
    # 可以使用中文路径读取图片
    cv_img = cv2.imdecode(np.fromfile(file_path, dtype=np.uint8), -1)
    return cv_img


device = 'cuda' if torch.cuda.is_available() else 'cpu'
print('Running inference on device \"{}\"'.format(device))
config = {
    'superpoint': {
        'nms_radius': 4,
        'keypoint_threshold': 0.001,
        'max_keypoints': -1
    },
    'superglue': {
        'weights': 'outdoor',
        'sinkhorn_iterations': 100,
        'match_threshold': 0.1,
    }
}
matching = Matching(config).eval().to(device)
matcher = LoFTR(config=default_cfg)
image_type = 'outdoor'
if image_type == 'indoor':
    matcher.load_state_dict(torch.load("loftr/weights/indoor_ds_new.ckpt")['state_dict'])
elif image_type == 'outdoor':
    matcher.load_state_dict(torch.load("loftr/weights/outdoor_ds.ckpt")['state_dict'])
else:
    raise ValueError("Wrong image_type is given.")
matcher = matcher.eval().to(device)


def nn_match_two_way(desc1, desc2, nn_thresh=0.7):
    if desc1.shape[1] == 0 or desc2.shape[1] == 0:
        return np.zeros((3, 0))
    if nn_thresh < 0.0:
        raise ValueError('\'nn_thresh\' should be non-negative')
    # Compute L2 distance. Easy since vectors are unit normalized.
    dmat = np.dot(desc1, desc2.T)
    dmat = np.sqrt(2 - 2 * np.clip(dmat, -1, 1))
    # Get NN indices and scores.
    idx = np.argmin(dmat, axis=1)
    scores = dmat[np.arange(dmat.shape[0]), idx]
    # Threshold the NN matches.
    keep = scores < nn_thresh
    # Check if nearest neighbor goes both directions and keep those.
    idx2 = np.argmin(dmat, axis=0)
    keep_bi = np.arange(len(idx)) == idx2[idx]
    keep = np.logical_and(keep, keep_bi)
    idx = idx[keep]
    scores = scores[keep]
    # Get the surviving point indices.
    m_idx1 = np.arange(desc1.shape[0])[keep]
    m_idx2 = idx
    # Populate the final 3xN match data structure.
    matches = np.zeros((3, int(keep.sum())))
    matches[0, :] = m_idx1
    matches[1, :] = m_idx2
    matches[2, :] = scores
    return matches


def sp_stitch_two(im11, im22):
    # print(im11.shape, im22.shape)
    im11 = cv2.cvtColor(im11, cv2.COLOR_BGR2GRAY)
    im22 = cv2.cvtColor(im22, cv2.COLOR_BGR2GRAY)
    im11 = cv2.resize(im11.astype('float32'), (W, H))
    im22 = cv2.resize(im22.astype('float32'), (W, H))

    inp0 = frame2tensor(im11, device)
    inp1 = frame2tensor(im22, device)
    pred, pred0, pred1 = matching({'image0': inp0, 'image1': inp1})
    # pred, pred0, pred1 = matching({'image0': inp0, 'image1': inp1, 'keypoints0': x1, 'keypoints1': x2, 'descriptors0':des1, 'descriptors1':des2})

    pred = {k: v[0].cpu().numpy() for k, v in pred.items()}
    kpts0, kpts1 = pred['keypoints0'], pred['keypoints1']
    # mdesc0, mdesc1 = pred['descriptors0'], pred['descriptors1']
    des_sp0, des_sp1 = (pred0['descriptors'][0].t()).cpu().numpy(), (pred1['descriptors'][0].t()).cpu().numpy()
    sp_matches = nn_match_two_way(des_sp0, des_sp1, nn_thresh=0.8)
    better = sp_matches[0:2, :]
    # for m, n, scores in sp_matches:
    #     better.append((m, n))
    # matches = pred['matches0']
    # confidence = pred['matching_scores0']
    _, num_match = better.shape
    mkpts0 = np.float32([kpts0[int(i)] for i in better[0, :]])
    mkpts1 = np.float32([kpts1[int(i)] for i in better[1, :]])
    # valid = matches > -1
    # mkpts0 = kpts0[valid]
    # mkpts1 = kpts1[matches[valid]]

    # print(len(mkpts0), len(mkpts1))
    if num_match > 10:
        (HH, status) = cv2.findHomography(mkpts1, mkpts0, cv2.RANSAC, ransacReprojThreshold=5.0)
        print(HH)
        if HH[0][0] <= 3 and HH[0][0] >= 0.4 and HH[1][1] <= 3 and HH[1][1] >= 0.4 and abs(HH[0][-1]) < W and abs(
                HH[1][-1]) < H:
            return True, HH, num_match
        return False, None, num_match
    else:
        return False, None, num_match


def sg_stitch_two(im11, im22):
    # print(im11.shape, im22.shape)
    im11 = cv2.cvtColor(im11, cv2.COLOR_BGR2GRAY)
    im22 = cv2.cvtColor(im22, cv2.COLOR_BGR2GRAY)
    im11 = cv2.resize(im11.astype('float32'), (W, H))
    im22 = cv2.resize(im22.astype('float32'), (W, H))

    inp0 = frame2tensor(im11, device)
    inp1 = frame2tensor(im22, device)
    pred, pred0, pred1 = matching({'image0': inp0, 'image1': inp1})
    # pred, pred0, pred1 = matching({'image0': inp0, 'image1': inp1, 'keypoints0': x1, 'keypoints1': x2, 'descriptors0':des1, 'descriptors1':des2})

    pred = {k: v[0].cpu().numpy() for k, v in pred.items()}
    kpts0, kpts1 = pred['keypoints0'], pred['keypoints1']
    mdesc0, mdesc1 = pred['descriptors0'], pred['descriptors1']
    matches = pred['matches0']
    confidence = pred['matching_scores0']
    valid = matches > -1
    mkpts0 = kpts0[valid]
    mkpts1 = kpts1[matches[valid]]

    # print(len(mkpts0), len(mkpts1))
    if len(mkpts0) > 10:
        (HH, status) = cv2.findHomography(mkpts1, mkpts0, cv2.RANSAC, ransacReprojThreshold=5.0)
        # print(HH)
        if HH[0][0] <= 3 and HH[0][0] >= 0.3 and HH[1][1] <= 3 and HH[1][1] >= 0.3 and abs(HH[0][-1]) < W and abs(
                HH[1][-1]) < H:
            return True, HH, len(mkpts0)
        return False, None, len(mkpts0)
    else:
        return False, None, len(mkpts0)


def loftr_stitch_two(im11, im22):
    # print(im11.shape, im22.shape)
    im11 = cv2.cvtColor(im11, cv2.COLOR_BGR2GRAY)
    im22 = cv2.cvtColor(im22, cv2.COLOR_BGR2GRAY)
    im11 = cv2.resize(im11.astype('float32'), (W, H))
    im22 = cv2.resize(im22.astype('float32'), (W, H))

    inp0 = frame2tensor(im11, device)
    inp1 = frame2tensor(im22, device)

    batch = {'image0': inp0, 'image1': inp1}
    with torch.no_grad():
        matcher(batch)
        mkpts0 = batch['mkpts0_f'].cpu().numpy()
        mkpts1 = batch['mkpts1_f'].cpu().numpy()
        mconf = batch['mconf'].cpu().numpy()
    print(len(mkpts0), len(mkpts1))
    if len(mkpts0) > 10:
        (HH, status) = cv2.findHomography(mkpts1, mkpts0, cv2.RANSAC, ransacReprojThreshold=5.0)
        print(HH)
        if HH[0][0] <= 3 and HH[0][0] >= 0.3 and HH[1][1] <= 3 and HH[1][1] >= 0.3 and abs(HH[0][-1]) < W and abs(
                HH[1][-1]) < H:
            return True, HH, len(mkpts0)
        return False, None, len(mkpts0)
    else:
        return False, None, len(mkpts0)


def sift_stitch_two(im11, im22):
    im11 = cv2.cvtColor(im11, cv2.COLOR_BGR2GRAY)
    im22 = cv2.cvtColor(im22, cv2.COLOR_BGR2GRAY)
    im11 = cv2.resize(im11.astype('uint8'), (W, H))
    im22 = cv2.resize(im22.astype('uint8'), (W, H))

    sift = cv2.SIFT_create()
    kp1, des1 = sift.detectAndCompute(im11, None)  # des是描述子
    kp2, des2 = sift.detectAndCompute(im22, None)  # des是描述子
    kp1 = np.float32([kp.pt for kp in kp1])
    kp2 = np.float32([kp.pt for kp in kp2])
    # print(des1, des2)
    if des1 is None or des2 is None:
        return False, None, 0
    matches = cv2.BFMatcher().knnMatch(des1, des2, k=2)
    good = []

    for m in matches:
        # 当最近距离跟次近距离的比值小于ratio值时，保留此匹配对
        if len(m) == 2 and m[0].distance < m[1].distance * 0.7:
            # 存储两个点在featuresA, featuresB中的索引值
            good.append((m[0].trainIdx, m[0].queryIdx))

    num_match = len(good)
    ptsA = np.float32([kp1[i] for (_, i) in good])
    ptsB = np.float32([kp2[i] for (i, _) in good])
    if num_match > 10:
        (HH, status) = cv2.findHomography(ptsB, ptsA, cv2.RANSAC, ransacReprojThreshold=5.0)

        print(HH)
        if HH[0][0] <= 3 and HH[0][0] >= 0.3 and HH[1][1] <= 3 and HH[1][1] >= 0.3 and abs(HH[0][-1]) < W and abs(
                HH[1][-1]) < H:
            return True, HH, num_match
        return False, None, num_match
    else:
        return False, None, num_match


def match_ratio(name1, name2):
    cnt = 0
    for i in range(len(name1)):
        if name1[i] != name2[i]:
            cnt += 1
    return cnt


def hh_stitch(image_list, HH_list, match_method):
    # print(len(image_list), len(HH_list))
    cur_index = len(HH_list)
    for i in range(cur_index + 1, len(image_list)):
        if match_method == 'sift':
            status, HH, num_match = sift_stitch_two(image_list[i - 1], image_list[i])
        if match_method == 'sp':
            status, HH, num_match = sp_stitch_two(image_list[i - 1], image_list[i])
        if match_method == 'sg':
            status, HH, num_match = sg_stitch_two(image_list[i - 1], image_list[i])
        if match_method == 'loftr':
            status, HH, num_match = loftr_stitch_two(image_list[i - 1], image_list[i])

        if not status:
            return 1, image_list, [], num_match
        HH_list.append(HH)

    return 0, image_list, HH_list, num_match
    # return 0, image_curr


def treat_list(image_list, homo_list):
    center = len(image_list) // 2
    # new_homo_list = [homo_list[center]]
    new_homo_list = []
    HH_bias = np.array([[1, 0, 0], [0, 1, 0], [0, 0, 1]])
    for i in range(0, center):
        HH_bias = np.dot(HH_bias, homo_list[i])
    HH_bias = np.linalg.inv(HH_bias)
    HH_cur_to_first = np.array([[1, 0, 0], [0, 1, 0], [0, 0, 1]])
    image_center = image_list[center]

    HH_use_list = []
    min_w = 1000
    min_h = 1000
    max_w = -1000
    max_h = -1000
    for i in range(len(image_list)):
        HH_cur = np.dot(HH_bias, HH_cur_to_first)
        f1 = np.dot(HH_cur, np.array([0, 0, 1]))
        f1 = f1 / f1[-1]
        f2 = np.dot(HH_cur, np.array([W, 0, 1]))
        f2 = f2 / f2[-1]
        f3 = np.dot(HH_cur, np.array([0, H, 1]))
        f3 = f3 / f3[-1]
        f4 = np.dot(HH_cur, np.array([W, H, 1]))
        f4 = f4 / f4[-1]
        print(i, HH_cur)
        HH_use = HH_cur
        HH_use_list.append(HH_use)
        min_w = min(min_w, min([f1[0], f2[0], f3[0], f4[0]]))
        min_h = min(min_h, min([f1[1], f2[1], f3[1], f4[1]]))
        max_w = max(max_w, max([f1[0], f2[0], f3[0], f4[0]]))
        max_h = max(max_h, max([f1[1], f2[1], f3[1], f4[1]]))
        if i < len(image_list) - 1:
            HH_cur_to_first = np.dot(HH_cur_to_first, homo_list[i])
    max_w_int = int(max_w + 1)
    max_h_int = int(max_h + 1)
    min_w_int = int(min_w - 1)
    min_h_int = int(min_h - 1)
    # print(max_w_int, max_h_int, min_w_int, min_h_int)
    newW = - min_w_int + max_w_int
    newH = - min_h_int + max_h_int
    # print(newW, newH)
    try:
        pano = np.zeros([newH, newW, 3], np.float32)
        image_cur_list = []
        for i in range(len(image_list)):
            HH_use = HH_use_list[i]
            HH_use[0][-1] += - min_w_int
            HH_use[1][-1] += - min_h_int
            image_cur = cv2.warpPerspective(image_list[i], HH_use, (newW, newH))
            image_cur_list.append(image_cur)
            # pano = pano + image_cur
            mask = image_cur != 0
            pano[mask] = image_cur[mask]
            # for ii in range(newH):
            #     for jj in range(newW):
            #         for cc in range(3):
            #             if image_cur[ii, jj, cc] != 0:
            #                 pano[ii, jj, cc] = image_cur[ii, jj, cc]
            if i < len(image_list) - 1:
                HH_cur_to_first = np.dot(HH_cur_to_first, homo_list[i])

        aa = np.sum(image_cur_list[0] * image_cur_list[1], axis=2) > 0
        # print(aa.shape, image_cur_list[0].shape, image_cur_list[1].shape)
        img1 = np.mean(image_cur_list[0][aa], axis=-1)
        img2 = np.mean(image_cur_list[1][aa], axis=-1)
        img1 = np.reshape(img1, -1)
        img2 = np.reshape(img2, -1)
        # print(img1.shape, img2.shape)
        MI = mutual_info_score(img1, img2)
        corr = np.corrcoef(img1, img2)[0, 1]
        # img11 = img1 / np.sum(img1)
        # img22 = img2 / np.sum(img2)
        hist1, _ = np.histogram(img1, bins=256, range=(0, np.max(img1)), density=True)
        hist2, _ = np.histogram(img2, bins=256, range=(0, np.max(img2)), density=True)
        # Compute entropy
        ent1 = entropy(hist1, base=2)
        ent2 = entropy(hist2, base=2)

        CE = ent1 + ent2 - MI
        # cross_entropy = -np.sum(img11 * np.log(img22))

        return pano, MI, corr, CE
    except:
        pano = np.zeros([W, H, 3], np.float32)
        img1 = np.mean(image_list[0], axis=-1)
        img2 = np.mean(image_list[1], axis=-1)
        hist1, _ = np.histogram(img1, bins=256, range=(0, 256), density=True)
        hist2, _ = np.histogram(img2, bins=256, range=(0, 256), density=True)
        # Compute entropy
        ent1 = entropy(hist1, base=2)
        ent2 = entropy(hist2, base=2)

        CE = ent1 + ent2

        return pano, 0, 0, CE


if __name__ == "__main__":
    print(cv2.__version__)
    # img1 = cv2.imread('dem\\demo2\\hw1.jpg')  # 图片绝对路径，
    # img2 = cv2.imread('dem\\demo2\\hw2.jpg')
    # img3 = cv2.imread('dem\\demo2\\hw3.jpg')
    # img4 = cv2.imread('dem\\st_123.jpg')

    parser = argparse.ArgumentParser()
    parser.add_argument('--dataset', default='kjg1', help='kjg1, kjg2, kjg3, kjg4, hw')
    parser.add_argument('--method', default='loftr', help='sift, sp, sg, loftr')
    parser.add_argument('--root', default='/home/fsh/stitching_and_registration/input_data/stitch_data/stitch')
    parser.add_argument('--outfile', default='/home/fsh/stitching_and_registration/output_data/stitch_out/stat')
    args = parser.parse_args()

    dataset = args.dataset
    match_method = args.method
    root = args.root
    out_root = args.outfile

    out_path = osp.join(out_root, dataset)

    if not osp.exists(out_path):
        os.makedirs(out_path)
    img_path = []

    max_row_sum = 2000

    img_path = []
    # base_path = 'F:\\hw_data\\鹏泉站智能巡视\\鹏泉站智能巡视\\1\\1\\室外巡检机器人可见光任务20220818145748'
    # excel_name = '室外巡检机器人可见光任务20220818145748.xlsx'

    if dataset == 'kjg1':
        base_path = osp.join(root, '(例)泉城变电站白天例行任务20220306070002')
        excel_name = '(例)泉城变电站白天例行任务20220306070002.xlsx'
        multi_list = [[129, 131, 133], [141, 142, 144, 146], [154, 156, 158], [172, 174, 175], [178, 179, 181],
                      [200, 201, 202, 203, 205, 207], [210, 212, 214, 217, 218], [229, 231, 233], [242, 243, 245, 246],
                      [308, 309, 311, 313], [321, 322, 323, 324, 326, 328]]
        start_line = 99
        image_col = 9
        name_col = 1
        part_col = 2
    if dataset == 'kjg2':
        base_path = osp.join(root, '(例)泉城变电站白天例行任务20220321070002')
        excel_name = '(例)泉城变电站白天例行任务20220321070002.xlsx'
        multi_list = [[129, 130, 131], [162, 164, 166, 168, 170], [202, 204, 205], [231, 232, 233, 234, 236, 238],
                      [241, 243, 245, 248, 249, 251],
                      [288, 290, 292], [297, 299, 301], [302, 305]]
        start_line = 129
        image_col = 9
        name_col = 1
        part_col = 2
    if dataset == 'kjg3':
        base_path = osp.join(root, '(例)泉城变电站白天例行任务20220428070002')
        excel_name = '(例)泉城变电站白天例行任务20220428070002.xlsx'
        multi_list = [[172, 174, 176], [190, 191, 193, 195], [244, 246, 248], [290, 291, 292],
                      [371, 372, 373, 374, 376, 378],
                      [393, 394, 396, 397], [414, 415, 418]]
        start_line = 148
        image_col = 9
        name_col = 1
        part_col = 2
    if dataset == 'kjg4':
        base_path = osp.join(root, '(例)泉城变电站白天例行任务20220501070002')
        excel_name = '(例)泉城变电站白天例行任务20220501070002.xlsx'
        multi_list = [[346, 347, 348], [438, 439, 440, 441], [164, 166, 168], [202, 203, 205, 207],
                      [204, 206, 208, 209],
                      [262, 263, 265, 267], [385, 386, 387, 388, 390, 392], [407, 408, 409, 410]]
        start_line = 160
        image_col = 9
        name_col = 1
        part_col = 2
        # print(os.path.exists('/data/fsh/pinjie_data/(例)泉城变电站白天例行任务20220501070002/image/110kV设备区1主变1121L低抗间隔112117接地刀闸A相本体刀闸外观20220501070702.jpg'))
    if dataset == 'hw':
        base_path = osp.join(root, '室外巡检机器人红外任务20220815172558')
        excel_name = '室外巡检机器人红外任务20220815172558.xlsx'
        multi_list = [[343, 344, 345], [346, 347, 348], [372, 373, 374, 375, 376, 377, 378],
                      [443, 444, 445, 446, 447, 448, 449, 450, 451],
                      [487, 488, 489], [502, 503, 504], [573, 574, 575, 578, 579], [600, 601, 602, 603],
                      [650, 651, 652], [665, 666, 667], [522, 523, 524, 525, 526, 527, 528, 529]]
        start_line = 342
        image_col = 10
        name_col = 4
        part_col = 1

    work_book = openpyxl.load_workbook(osp.join(base_path, excel_name))
    sheet_1 = work_book.active
    # print(sheet_1.dimensions)
    row_sum = sheet_1.max_row
    row_0_value = sheet_1.cell(15, 6)
    # print(row_sum, row_0_value)
    name_last = ''
    part_last = ''
    stitcher = cv2.Stitcher.create(cv2.Stitcher_PANORAMA)
    image_list = []
    notfirst = 0
    cnt = 0
    cnt_pinjie = 0
    MI_list = []
    num_match_list = []
    index_list = []
    name_list = []

    corr_list = []
    ce_list = []

    total = 0
    success = 0
    # start_line = 500
    for i in range(start_line, start_line + max_row_sum):
        # print(i)
        # name = sheet_1.cell(i, 1).value[:-2]
        name = sheet_1.cell(i, name_col).value
        part = sheet_1.cell(i, part_col).value
        # print(name, name_last)
        # if name!=name_last and notfirst:
        # print(part, part_last)

        if name is None or name_last is None: continue
        if len(name) != len(name_last) or part != part_last or match_ratio(name, name_last) > 2:
            image_list = []
            index_list = []
            name_list = []
            homo_list = []
            part_list = []

        url_map = sheet_1.cell(i, image_col).value
        # print(url_map)
        url = sheet_1.cell(i, image_col).hyperlink.target
        # print(url)
        image_path = osp.join(osp.join(base_path, 'image'), url[6:])
        try:
            img1 = cv_imread(image_path)
        except:
            print(i)
            image_list = []
            index_list = []
            name_list = []
            homo_list = []
            part_list = []
            name_last = ''
            part_last = ''
            continue
        img1 = cv2.resize(img1.astype('float32'), (W, H))
        img1_gray = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY)
        image_list.append(img1)
        index_list.append(i)
        name_list.append(name)
        part_list.append(part)
        notfirst = 1

        if len(image_list) > 2:
            image_list = image_list[1:]
            index_list = index_list[1:]
            name_list = name_list[1:]
            homo_list = homo_list[1:]
            part_list = part_list[1:]
        if len(image_list) == 2:
            # image_list = []
            print(name_list, part_list)
            cnt += 1

            (status, image_list, homo_list, num_match) = hh_stitch(image_list, homo_list, match_method)
            num_match_list.append((num_match))
            print('num match means:')
            print(sum(num_match_list) / len(num_match_list), len(num_match_list))
            # s = MyStitch(image_list, homo_list)
            # s.leftshift()
            # s.rightshift()
            # pano = s.leftImage
            # print(status)
            if status == cv2.Stitcher_OK:

                pano, MI, corr, CE = treat_list(image_list, homo_list)

                MI_list.append(MI)
                corr_list.append(corr)
                ce_list.append(CE)

                print('MI means:')
                print(MI, sum(MI_list) / len(MI_list), len(MI_list))
                print('corr means:')
                print(corr, sum(corr_list) / len(corr_list), len(corr_list))
                print('CE means:')
                print(CE, sum(ce_list) / len(ce_list), len(ce_list))
                print('Success rate:')
                print(cnt_pinjie / len(num_match_list), len(MI_list), len(num_match_list))
                cnt_pinjie += 1
                print(cnt_pinjie, cnt)
                print("拼接成功.")


            else:
                img1 = np.mean(image_list[0], axis=-1)
                img2 = np.mean(image_list[1], axis=-1)
                hist1, _ = np.histogram(img1, bins=256, range=(0, 256), density=True)
                hist2, _ = np.histogram(img2, bins=256, range=(0, 256), density=True)
                # Compute entropy
                ent1 = entropy(hist1, base=2)
                ent2 = entropy(hist2, base=2)
                CE = ent1 + ent2
                MI_list.append(0)
                corr_list.append(0)
                ce_list.append(CE)
                image_list = image_list[1:]
                index_list = index_list[1:]
                name_list = name_list[1:]
                homo_list = homo_list[1:]
                part_list = part_list[1:]
                # image_list.pop(0)
            # MI_list.append(0)
            # num_match_list.append(0)
            # cnt += 1
            image_list = image_list[1:]
            index_list = index_list[1:]
            name_list = name_list[1:]
            homo_list = homo_list[1:]
            part_list = part_list[1:]
        name_last = name
        part_last = part
    print(cnt_pinjie, cnt)
    print('Success rate:')
    print(cnt_pinjie / len(num_match_list), len(MI_list), len(num_match_list))
    print('MI means:')
    print(sum(MI_list) / len(MI_list), len(MI_list))
    print('corr means:')
    print(sum(corr_list) / len(corr_list), len(corr_list))
    print('CE means:')
    print(sum(ce_list) / len(ce_list), len(ce_list))
    print('num match means:')
    print(sum(num_match_list) / len(num_match_list), len(num_match_list))
    MI_npy = np.array(MI_list)
    num_match_npy = np.array(num_match_list)

    corr_npy = np.array(corr_list)
    ce_npy = np.array(ce_list)

    np_save_mi_path = out_path + '/MI' + '_' + dataset + '_' + match_method + '.npy'
    np_save_num_match_path = out_path + '/num_match' + '_' + dataset + '_' + match_method + '.npy'
    np.save(np_save_mi_path, MI_npy)
    np.save(np_save_num_match_path, num_match_npy)

    np_save_ce_path = out_path + '/CE' + '_' + dataset + '_' + match_method + '.npy'
    np_save_corr_path = out_path + '/corr' + '_' + dataset + '_' + match_method + '.npy'
    np.save(np_save_ce_path, ce_npy)
    np.save(np_save_corr_path, corr_npy)
